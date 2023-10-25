from typing import List, Tuple, Iterator, Union, Callable, Any
import enum

import dataclasses



class Token_Kind(enum.Enum):
    INT = enum.auto(),
    FLOAT = enum.auto(),
    STRING = enum.auto(),
    PLUS = enum.auto(),
    MINUS = enum.auto(),
    ASTERISK = enum.auto(),
    SLASH = enum.auto(),
    LPAREN = enum.auto(),
    RPAREN = enum.auto(),
    FUNC = enum.auto(),
    COMMA = enum.auto(),
    REF = enum.auto(),
    RANGE = enum.auto(),


@dataclasses.dataclass
class Token:
    kind: Token_Kind
    value: Union[int, str]

    def __str__(self):
        return f"Token(kind={self.kind:20}, value={self.value})"


def match_sequence_in_range(s: str, rs: Iterator[Iterator]) -> Tuple[str, str]:
    out = ""
    while s and any(ord(s[0]) in r for r in rs):
        c, s = s[:1], s[1:]
        out += c
    return out, s


def match_char_ranges(s: str, rs: Iterator[Iterator]):
    c = ""
    if s and any(ord(s[0]) in r for r in rs):
        c, s = s[:1], s[1:]
    return c, s


def match_digits(s: str)-> Tuple[str,str]:
    DIGITS = [range(ord("0"), ord("9")+1)]
    return match_sequence_in_range(s, DIGITS)

def match_alphas_upper(s:str) -> Tuple[str,str]:
    LOWERS = [range(ord("A"),ord("Z")+1)]
    return match_sequence_in_range(s,LOWERS)


def match_float(s:str) -> Tuple[str,str]:
    start = s
    flt_int, s = match_digits(s)
    if not flt_int:
        return "", start
    
    if not (s and s[0] == "."):
        return "",start

    _, s = s[:1], s[1:]

    flt_frac, s = match_digits(s) # this can be empty, as in '1.'

    flt_str = f"{flt_int}.{flt_frac}"
    return flt_str, s


def match_whitespace(s: str):
    WHITES = [[ord(c) for c in [" ", "\t"]]]
    return match_sequence_in_range(s, WHITES)

def match_escape_sequence(s:str):
    start = s
    if not (s and s[0] == "\\"):
        return "",start
    _, s = s[:1], s[1:]
    c, s = s[:1], s[1:]
    if c == "\\":
        return "\\",s
    elif c == "n":
        return "\n",s
    elif c == "t":
        return "\t",s
    elif c =="\"":
        return "\""
    elif c == "\'":
        return "\'"
    else:
        assert False, f"undefined escape sequence: '\{c}'"


def match_string(s:str):
    str_seps = ["\"", "'"]

    start = s
    if not (s and s[0] in str_seps):
        return None,start

    str_sep,s = s[:1], s[1:]

    str_text = ""
    while True:
        if len(s) == 0:
            return None,start
        elif s[0] == str_sep:
            _,s = s[:1], s[1:]
            return str_text,s
        elif s[0] == "\\":
            esc,s = match_escape_sequence(s)
            str_text += esc
            continue
        else:
            c,s = s[:1], s[1:]
            str_text += c


def match_funcs(s: str):
    FUNCS = ["sq", "if","type"]
    s_low = s.lower()
    for func in FUNCS:
        if s_low.startswith(func):
            func_len = len(func)
            return s[:func_len], s[func_len:]
    return "", s

def match_ref(s:str) -> Tuple[str,str]:
    start = s
    if s.startswith("$"):
        _,s = s[:1],s[1:]
    rowstr,s = match_alphas_upper(s)
    if not rowstr:
        return "", start
    if s.startswith("$"):
        _,s = s[:1],s[1:]
    colstr,s = match_digits(s)
    if not colstr:
        return "", start
    refstr = rowstr+colstr
    return refstr,s


def match_range(s:str) -> Tuple [str,str]:
    start = s
    range_start, s = match_ref(s)
    if not range_start:
        return "",start
    if not (s and s[0] == ":"):
        return "",start
    _,s = s[:1],s[1:]
    range_end,s = match_range(s)
    if not range_start:
        return "",start
    range_str = f"{range_start}:{range_end}"
    return range_str,s


def tokenize(s: str) -> List[Token]:
    assert s, "Empty Input"
    start = s
    tokens = []
    while s:
        if s[0] == "+":
            c, s = s[:1], s[1:]
            tokens.append(Token(Token_Kind.PLUS, c))
            _, s = match_whitespace(s)
            continue
        elif s[0] == "-":
            c, s = s[:1], s[1:]
            tokens.append(Token(Token_Kind.MINUS, c))
            _, s = match_whitespace(s)
            continue
        elif s[0] == "*":
            c, s = s[:1], s[1:]
            tokens.append(Token(Token_Kind.ASTERISK, c))
            _, s = match_whitespace(s)
            continue
        elif s[0] == "/":
            c, s = s[:1], s[1:]
            tokens.append(Token(Token_Kind.SLASH, c))
            _, s = match_whitespace(s)
            continue
        elif s[0] == "(":
            c, s = s[:1], s[1:]
            tokens.append(Token(Token_Kind.LPAREN, c))
            _, s = match_whitespace(s)
            continue
        elif s[0] == ")":
            c, s = s[:1], s[1:]
            tokens.append(Token(Token_Kind.RPAREN, c))
            _, s = match_whitespace(s)
            continue
        elif s[0] == ",":
            c, s = s[:1], s[1:]
            tokens.append(Token(Token_Kind.COMMA, c))
            _, s = match_whitespace(s)

        str_str, s = match_string(s)
        if str_str is not None:
            tokens.append(Token(Token_Kind.STRING,str_str))
            _,s = match_whitespace(s)
            continue

        fun_str, s = match_funcs(s)
        if fun_str:
            tokens.append(Token(Token_Kind.FUNC, fun_str))
            _, s = match_whitespace(s)
            continue

        flt_str, s = match_float(s)
        if flt_str:
            tokens.append(Token(Token_Kind.FLOAT, float(flt_str)))
            _, s = match_whitespace(s)
            continue

        int_str, s = match_digits(s)
        if int_str:
            tokens.append(Token(Token_Kind.INT, int(int_str)))
            _, s = match_whitespace(s)
            continue

        ref_str, s = match_ref(s)
        if ref_str:
            tokens.append(Token(Token_Kind.REF, ref_str))
            _, s = match_whitespace(s)
            continue

        range_str, s = match_range(s)
        if range_str:
            tokens.append(Token(Token_Kind.RANGE,ref_str))
            _, s = match_whitespace(s)
            continue

        err_pos = len(start) - len(s)
        raise NotImplementedError(
            f"Unknown sequence at row {err_pos}:\n"
            f"{start}\n"
            f"{' '*err_pos}^"
        )

    return tokens


def validate(tokens: List[Token]):

    def validate_scopes(tokens: List[Token])->bool:
        depth = 0
        for token in tokens:
            if token.kind == Token_Kind.LPAREN:
                depth += 1
            elif token.kind == Token_Kind.RPAREN:
                depth -= 1
            
            if depth < 0:
                print("Unmatched right parenthesis")
                return False
        
        if depth != 0:
            print("Unmatched left parenthesis")
            return False
        return True
    
    def validate_grammar(tokens:List[Token])->bool:
        # grammar in this case refers to the order of literals <lit> and operators <op>.

        GRAMMAR_LIT = 0
        GRAMMAR_OP = 1
        structure = []

        kinds = [token.kind for token in tokens]
        for i,kind in enumerate(kinds):
            if kind in VAL_KINDS:
                structure.append(GRAMMAR_LIT)
            elif kind == Token_Kind.FUNC:
                if not i+1 < len(kinds) or kinds[i+1] != Token_Kind.LPAREN:
                    print(f"Function '{tokens[i].value}' does not have an argument list.")
                    return False
                structure.append(GRAMMAR_LIT)
                structure.append(GRAMMAR_OP)
            elif kind == Token_Kind.LPAREN:   # '(' -> <lit>, <op> maintains the grammar structure
                structure.append(GRAMMAR_LIT)
                structure.append(GRAMMAR_OP)
            elif kind == Token_Kind.RPAREN:   # ')' ->  <op>, <lit> maintains the grammar structure
                if i-2 >= 0 and kinds[i-2:i] == [Token_Kind.FUNC, Token_Kind.LPAREN]:
                    # handle edge case 'f()'
                    structure.append(GRAMMAR_LIT)
                structure.append(GRAMMAR_OP)
                structure.append(GRAMMAR_LIT)
            elif kind in OP_KINDS:
                structure.append(GRAMMAR_OP)
            else:
                assert False, f"Unknown grammar token {tokens[i]}"
        
        # a valid grammar follows the pattern <lit> <op> <lit> <op> ... <op> <lit>,
        # alternating literals and operators, and starting and ending with a literal.

        if len(structure)%2 == 0:
            print("Invalid syntax")
            return False

        for i,elem in enumerate(structure):
            if not elem == i%2:
                print("Invalid syntax")
                return False
        return True
    
    def validate_functions(tokens:List[Token]):

        FUNC_OPCOUNT = {
            "sq":1,
            "if":3,
            "type":1,
        }
        def match_scope(tokens:List[Token]):
            if not (tokens and tokens[0].kind == Token_Kind.LPAREN):
                return [],tokens

            depth = 1
             
            for i,token in enumerate(tokens[1:]):
                if token.kind == Token_Kind.LPAREN:
                    depth += 1
                elif token.kind == Token_Kind.RPAREN:
                    depth -= 1
                
                if depth == 0:
                    return tokens[:i+2],tokens[i+2:]
            return [], tokens

        def split_exprlist(tokens:List[Token]):
            lp, *exprl, rp = tokens
            assert lp.kind == Token_Kind.LPAREN
            assert rp.kind == Token_Kind.RPAREN
            if len(exprl) == 0:
                return []
            
            exprs = []
            expr = list()
            depth = 0
            for token in exprl:
                if token.kind == Token_Kind.LPAREN:
                    depth += 1
                    expr.append(token)
                elif token.kind == Token_Kind.RPAREN:
                    depth -= 1
                    expr.append(token)
                elif token.kind == Token_Kind.COMMA:
                    if depth == 0:
                        exprs.append(expr)
                        expr = list()
                    else:
                        expr.append(token)
                else:
                    expr.append(token)
            
            exprs.append(expr)
            return exprs
        
        while tokens:
            token = tokens[0]
            if token.kind == Token_Kind.FUNC:
                func, *tokens = tokens
                exprl, tokens = match_scope(tokens)
                exprs = split_exprlist(exprl)
                if not FUNC_OPCOUNT[func.value] == len(exprs):
                    plural = "argument" if FUNC_OPCOUNT[func.value] == 1 else "arguments"
                    print(f"'{func.value}' expects {FUNC_OPCOUNT[func.value]} {plural} but got {len(exprs)}")
                    return False
                
                if not all(validate_functions(expr) for expr in exprs):
                    return False

            else:
                _, *tokens = tokens
                
        return True


    tests=[
        validate_scopes,
        validate_grammar,
        validate_functions,
    ]

    for test in tests:
        if not test(tokens):
            return False
    return True
    

LIT_KINDS = [
    Token_Kind.INT,
    Token_Kind.FLOAT,
    Token_Kind.STRING,
]

VAR_KINDS = [
    Token_Kind.REF,
    Token_Kind.RANGE,
]

VAL_KINDS = LIT_KINDS + VAR_KINDS

OP_PREC = {
    Token_Kind.COMMA: 0, # TODO: compare this to precedence 4
    Token_Kind.PLUS: 1,
    Token_Kind.MINUS: 1,
    Token_Kind.ASTERISK: 2,
    Token_Kind.SLASH: 2,
    Token_Kind.FUNC: 3,
    Token_Kind.LPAREN: 4,
}

OP_KINDS = [
    Token_Kind.PLUS,
    Token_Kind.MINUS,
    Token_Kind.ASTERISK,
    Token_Kind.SLASH,
    Token_Kind.COMMA,
    Token_Kind.FUNC,
    Token_Kind.LPAREN,
]


def parse(tokens: List[Token]) -> List[Token]:
    '''Translate infix notation to postfix notation.'''

    def peek(stack):
        assert stack, "Peek from empty stack."
        return stack[-1]


    ops: List[Token] = []
    out: List[Token] = []

    for token in tokens:
        if token.kind in VAL_KINDS:
            out.append(token)
        elif token.kind in OP_KINDS:
            while ops and peek(ops).kind not in [Token_Kind.LPAREN, Token_Kind.COMMA] and not OP_PREC[peek(ops).kind] < OP_PREC[token.kind]:
                out.append(ops.pop())
            ops.append(token)
        elif token.kind == Token_Kind.RPAREN:
            while ops and peek(ops).kind != Token_Kind.LPAREN:
                if peek(ops).kind == Token_Kind.COMMA:
                    _ = ops.pop()
                else:
                    out.append(ops.pop())
            if ops and peek(ops).kind == Token_Kind.FUNC:
                out.append(ops.pop())
            assert ops.pop().kind == Token_Kind.LPAREN
        else:
            raise NotImplementedError(f"Unparsable token: {token}")

    while ops:
        out.append(ops.pop())
    return out


def make_token(a:Union[int,float,str])-> Token:
    # piggyback off of python's type system... for now.
    if isinstance(a, int):
        return Token(Token_Kind.INT,a)
    elif isinstance(a, float):
        return Token(Token_Kind.FLOAT,a)
    elif isinstance(a,str):
        return Token(Token_Kind.STRING,a)
    else:
        assert False, f"Unknown type '{type(a)}' of value '{a}'"


def evaluate(tokens: List[Token],lookup_callback:Callable[[str],Any]):
    stack: List[Token] = []
    for token in tokens:
        if token.kind in LIT_KINDS:
            stack.append(token)
        elif token.kind == Token_Kind.REF:
            a = lookup_callback(token.value)
            stack.append(make_token(a))
        elif token.kind == Token_Kind.RANGE:
            raise NotImplementedError("Evaluating ranges is not implemented yet.")
        elif token.kind == Token_Kind.PLUS:
            b = stack.pop().value
            a = stack.pop().value
            c = a+b
            stack.append(make_token(c))
        elif token.kind == Token_Kind.MINUS:
            b = stack.pop().value
            a = stack.pop().value
            c = a-b
            stack.append(make_token(c))
        elif token.kind == Token_Kind.ASTERISK:
            b = stack.pop().value
            a = stack.pop().value
            c = a*b
            stack.append(make_token(c))
        elif token.kind == Token_Kind.SLASH:
            b = stack.pop().value
            a = stack.pop().value
            c = a/b    
            stack.append(make_token(c))
        elif token.kind == Token_Kind.FUNC:
            if token.value == "sq":
                assert not len(stack) < 1, "Not enough arguments for sq function"
                a = stack.pop().value
                b = a*a
                stack.append(make_token(b))
            elif token.value == "if":
                assert not len(stack) < 3, "Not enough arguments for if function"
                false_val = stack.pop().value
                true_val = stack.pop().value
                condition = stack.pop().value
                if condition:
                    stack.append(make_token(true_val))
                else:
                    stack.append(make_token(false_val))
            elif token.value == "type":
                assert not len(stack) < 1, "Not enough arguments for type function"
                a = stack.pop().kind
                stack.append(Token(Token_Kind.STRING,a))
            else:
                raise NotImplementedError(
                    f"Cannot evaluate unknown function: {token}")
        else:
            raise NotImplementedError(
                f"Cannot evaluate unknown token: {token}")
    assert len(stack) == 1
    return stack.pop()


def calculate(s: str, lookup_callback:Callable[[str],Any]):
    tokens = tokenize(s)
    if not validate(tokens): return None
    instructions = parse(tokens)
    result = evaluate(instructions,lookup_callback)
    return result.value


def print_tokens(tokens:List[Token]):
    print(" ".join(str(token.value) for token in tokens))

def terminal(lookup_callback:Callable[[str],Any]):
    while True:
        r = calculate(input("> "),lookup_callback)
        if r is not None:
            print(r)

def xl_terminal(lookup_callback:Callable[[str],Any]):
    while True:
        r = xl_evalref(input("> ",lookup_callback))
        if r is not None:
            print(r)


def xl_evalref(lookup_callback:Callable[[str],Any],ref:str):
    value = ws[ref].value
    if isinstance(value,str) and value.startswith("="):
        result = calculate(value[1:],lookup_callback)
        return result
    else:
        return value


if __name__ == "__main__":
    import openpyxl as xl

    wb = xl.load_workbook("XL_TEST.xlsx")
    ws = wb.active

    result = xl_evalref(lambda ref: xl_evalref(ws[ref].value,ref),"C1")
    print(result)
    # terminal()
    # source = "1    +3*9*((7) + 3) "
    # source = "sq(sq(2))"
    # source = "sq(if(1,10*1/1+1-1,20))"
    # tokens = tokenize(source)  # 271
    # tokens = tokenize("1 + sq(2)")
    # tokens = tokenize("1    +3*9*sq((7) + 3) ") # 2701
    # tokens = tokenize("10+if(1,sq(if(1,1000,5)),10)") # 1000010
    # tokens = tokenize("if(1,if(if(0,0,1),2,3),4)") # =2

    # print_tokens(tokens)
    # # for token in tokens:
    # #     print(token)

    # instr = parse(tokens)
    # print()
    # print_tokens(instr)

    # result = evaluate(instr).value
    # print(f"{result=}")

    # print(calculate(source))




